# -*- coding: utf-8 -*-
"""
extract_xlsx.py
================

Extrae del archivo Excel (Plantilla-diagnóstico-financiero-(Ejemplo).xlsx)
toda la información legible (valores, fórmulas, estructura, dependencias)
y la imprime en consola como un resumen ejecutivo.

El script está pensado como punto de partida para un análisis manual más
profundo.  Genera un documento Markdown aparte (`analisis-plantilla-financiera.md`)
que ya queda persistido en `docs/`.

Reglas de la extracción:
- Carga el archivo dos veces:
    * Una con `data_only=False` para conservar las fórmulas.
    * Otra con `data_only=True` para comparar valores calculados y
      detectar celdas con error (#REF!, #DIV/0!, #VALUE!, #NAME?, etc.).
- Para cada hoja detecta el rango con datos (fila/columna mínima/máxima).
- Extrae celdas con valores literales, celdas con fórmulas y referencias
  inter-hoja.

Uso:
    python scripts/extract_xlsx.py
"""

from __future__ import annotations

import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "docs" / "Plantilla-diagnóstico-financiero-(Ejemplo).xlsx"

# Patrones útiles para detectar referencias inter-hoja en fórmulas
SHEET_REF_PATTERN = re.compile(
    r"(?:'?([^'!]+)'?\!|([A-Za-z_][\w\s\.]*)\!)",  # 'Hoja'!A1  o  Hoja!A1
    flags=re.IGNORECASE,
)
EXCEL_ERROR_VALUES = {
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
}
FORMULA_TRUNCATE_LEN = 200  # En el .md se truncan a partir de este largo


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------


def is_formula(cell: Cell) -> bool:
    """Devuelve True si la celda contiene una fórmula."""
    return isinstance(cell.value, str) and cell.value.startswith("=")


def cell_repr(cell: Cell) -> str:
    """Representación corta del valor de la celda para usar en tablas."""
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, float):
        # Evitar notación científica fea para valores financieros
        if abs(val) >= 1:
            return f"{val:,.2f}"
        return f"{val:.6f}"
    s = str(val)
    if len(s) > 80:
        return s[:77] + "..."
    return s


def detect_used_range(ws: Worksheet) -> tuple[int, int, int, int] | None:
    """Devuelve (min_row, max_row, min_col, max_col) del rango con datos.
    Devuelve None si la hoja está completamente vacía.
    """
    min_row, max_row = None, None
    min_col, max_col = None, None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if min_row is None or cell.row < min_row:
                min_row = cell.row
            if max_row is None or cell.row > max_row:
                max_row = cell.row
            if min_col is None or cell.column < min_col:
                min_col = cell.column
            if max_col is None or cell.column > max_col:
                max_col = cell.column
    if min_row is None:
        return None
    return min_row, max_row, min_col, max_col


def extract_inter_sheet_refs(formula: str, current_sheet: str) -> list[str]:
    """Devuelve la lista (única) de hojas referenciadas por una fórmula."""
    refs: set[str] = set()
    # Patrones: 'Nombre hoja'!A1  o  NombreHoja!A1
    for match in re.finditer(r"'([^']+)'!", formula):
        sheet_name = match.group(1).strip()
        if sheet_name and sheet_name != current_sheet:
            refs.add(sheet_name)
    for match in re.finditer(r"\b([A-Za-z_][\w\s\.]*?)\!", formula):
        sheet_name = match.group(1).strip()
        if (
            sheet_name
            and sheet_name != current_sheet
            and not sheet_name.isdigit()  # excluye referencias tipo 1!A1 que no aplican acá
        ):
            refs.add(sheet_name)
    return sorted(refs)


# ---------------------------------------------------------------------------
# Extracción por hoja
# ---------------------------------------------------------------------------


def analyze_sheet(
    ws_formulas: Worksheet, ws_values: Worksheet
) -> dict[str, Any]:
    """Devuelve un dict con la información extraída de una hoja."""

    sheet_name = ws_formulas.title
    used = detect_used_range(ws_formulas)
    if used is None:
        return {
            "name": sheet_name,
            "empty": True,
            "range_str": "",
            "literal_cells": 0,
            "formula_cells": 0,
            "literal_values": [],
            "formulas": [],
            "inter_sheet_refs": [],
            "errors": [],
        }

    min_row, max_row, min_col, max_col = used
    range_str = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )

    literal_values: list[dict[str, Any]] = []
    formulas: list[dict[str, Any]] = []
    inter_sheet_refs: set[str] = set()
    errors: list[dict[str, Any]] = []

    for row in ws_formulas.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            coord = cell.coordinate
            if cell.value is None:
                continue
            if is_formula(cell):
                formula_text = cell.value  # type: ignore[assignment]
                # Valor calculado correspondiente
                calc_cell = ws_values[coord]
                calc_val = calc_cell.value
                error_kind = None
                if isinstance(calc_val, str) and calc_val in EXCEL_ERROR_VALUES:
                    error_kind = calc_val
                    errors.append(
                        {
                            "coord": coord,
                            "formula": formula_text,
                            "error": error_kind,
                        }
                    )
                # Detectar refs inter-hoja
                for ref in extract_inter_sheet_refs(formula_text, sheet_name):
                    inter_sheet_refs.add(ref)
                formulas.append(
                    {
                        "coord": coord,
                        "formula": formula_text,
                        "calculated": calc_val,
                        "error": error_kind,
                    }
                )
            else:
                literal_values.append(
                    {
                        "coord": coord,
                        "value": cell.value,
                    }
                )

    return {
        "name": sheet_name,
        "empty": False,
        "range_str": range_str,
        "literal_cells": len(literal_values),
        "formula_cells": len(formulas),
        "literal_values": literal_values,
        "formulas": formulas,
        "inter_sheet_refs": sorted(inter_sheet_refs),
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# Resumen ejecutivo
# ---------------------------------------------------------------------------


def print_executive_summary(sheets_info: list[dict[str, Any]], wb: Workbook) -> None:
    """Imprime un resumen ejecutivo del Excel en consola."""
    total_sheets = len(sheets_info)
    total_literals = sum(s["literal_cells"] for s in sheets_info)
    total_formulas = sum(s["formula_cells"] for s in sheets_info)
    total_errors = sum(len(s["errors"]) for s in sheets_info)
    file_size = EXCEL_PATH.stat().st_size if EXCEL_PATH.exists() else 0

    print("=" * 72)
    print("RESUMEN EJECUTIVO — Plantilla diagnóstico financiero (Ejemplo).xlsx")
    print("=" * 72)
    print(f"Archivo           : {EXCEL_PATH.name}")
    print(f"Ruta absoluta     : {EXCEL_PATH}")
    print(f"Tamaño            : {file_size:,} bytes "
          f"({file_size / 1024:.1f} KiB)")
    print(f"Total de hojas    : {total_sheets}")
    print(f"Celdas con valor  : {total_literals:,}")
    print(f"Celdas con fórmula: {total_formulas:,}")
    print(f"Total celdas activas: {total_literals + total_formulas:,}")
    print(f"Errores detectados: {total_errors}")
    print()
    print("Inventario de hojas:")
    print("-" * 72)
    print(f"{'#':<3} {'Hoja':<35} {'Rango':<14} {'Lit':>5} {'Fórm':>5} {'Refs':>5}")
    print("-" * 72)
    for i, s in enumerate(sheets_info, 1):
        print(
            f"{i:<3} {s['name'][:34]:<35} "
            f"{s['range_str'][:13]:<14} "
            f"{s['literal_cells']:>5} "
            f"{s['formula_cells']:>5} "
            f"{len(s['inter_sheet_refs']):>5}"
        )
    print("-" * 72)
    print()
    print("Dependencias inter-hoja detectadas:")
    print("-" * 72)
    for s in sheets_info:
        if s["inter_sheet_refs"]:
            print(f"  {s['name']} -> {', '.join(s['inter_sheet_refs'])}")
    if not any(s["inter_sheet_refs"] for s in sheets_info):
        print("  (no se detectaron referencias inter-hoja explícitas)")
    print()
    if total_errors:
        print("Errores detectados:")
        print("-" * 72)
        for s in sheets_info:
            for err in s["errors"]:
                print(f"  {s['name']}!{err['coord']}  "
                      f"-> {err['error']}  ({err['formula']})")
    else:
        print("Errores detectados: ninguno (las fórmulas se evaluan sin error).")
    print()
    print("=" * 72)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    if not EXCEL_PATH.exists():
        print(f"ERROR: no se encuentra el archivo en {EXCEL_PATH}", file=sys.stderr)
        return 1

    # Carga con fórmulas
    wb_f = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    # Carga con valores calculados
    wb_v = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    sheets_info: list[dict[str, Any]] = []
    for name in wb_f.sheetnames:
        ws_f = wb_f[name]
        # Si la hoja no existe en wb_v (raro), usar la misma referencia
        ws_v = wb_v[name] if name in wb_v.sheetnames else wb_f[name]
        sheets_info.append(analyze_sheet(ws_f, ws_v))

    print_executive_summary(sheets_info, wb_f)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
